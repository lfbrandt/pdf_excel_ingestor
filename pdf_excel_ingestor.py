# pdf_excel_ingestor.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF→Excel Ingestor — robusto
- 1 página = 1 registro
- Captura dedicada de 'Mãe:'
- Correção de homoglifos (grego/cirílico)
- OCR por página automático se detectar glifos estranhos ou com --force-ocr
- OCR por página via ocrmypdf; fallback via pdf2image+pytesseract
- Dedupe contra XLSX, append no mesmo arquivo, relatórios

Atualizações:
- Remoção de espaços invisíveis (NBSP, thin, zero-width, etc.) em todos os textos
- Escrita no Excel com formato Texto (@) nas colunas sensíveis (IDs/datas)
- Alinhamento à esquerda e recuo = 0 em TODAS as colunas (COM e openpyxl).
- Colunas de data ("nascimento" e "data_admissao") tratadas como TEXTO (@) para manter alinhadas à esquerda.

Novidades desta versão:
- Força fonte Calibri em TODAS as células gravadas (COM e openpyxl) para evitar “nomes em grego” por estilo de coluna herdado.
- Mais invisíveis suportados: BOM/Zero-width no-break / soft hyphen.
- SOS S/N desativado (nunca preenche).
- Amarelo somente nas CÉLULAS VAZIAS das linhas gravadas (COM/openpyxl e writer simples).
"""

from __future__ import annotations
import argparse, csv, json, logging, re, sys, unicodedata, yaml, pdfplumber
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from glob import glob
from dateutil import parser as dateparser
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent
LOG = logging.getLogger("pdf_excel_ingestor")
LOG_FORMAT = "%(asctime)s | %(levelname)-5s | %(message)s"

# ----------------------------- Utils -----------------------------------------
def strip_accents_lower(s: str) -> str:
    if s is None:
        return ""
    s_nfkd = unicodedata.normalize("NFKD", s)
    return "".join([c for c in s_nfkd if not unicodedata.combining(c)]).lower()

def find_first(pattern: str, text: str) -> Optional[str]:
    m = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
    return m.group(0) if m else None

def clean_digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

# --- Invisíveis / whitespace --------------------------------------------------
INVISIBLE_SPACES = [
    "\u00A0",  # NBSP
    "\u202F",  # NNBSP
    "\u2007",  # Figure space
    "\u2009",  # Thin space
    "\u200A",  # Hair space
    "\u200B",  # Zero-width space
    "\u200C",  # ZW non-joiner
    "\u200D",  # ZW joiner
    "\u2060",  # Word joiner
    "\ufeff",  # BOM
    "\u2063",  # INVISIBLE SEPARATOR
    "\u00ad",  # SOFT HYPHEN
]

def normalize_ws(s: Optional[str]) -> Optional[str]:
    if not isinstance(s, str):
        return s
    for ch in INVISIBLE_SPACES:
        s = s.replace(ch, " ")
    s = re.sub(r"[ \t\r\f\v]+", " ", s)
    return s.strip()

# --- Homoglyph fix -----------------------------------------------------------
HOMOGLYPH_MAP = {
    # Grego
    "Α":"A","Β":"B","Ε":"E","Ζ":"Z","Η":"H","Ι":"I","Κ":"K","Μ":"M","Ν":"N","Ο":"O","Ρ":"R","Τ":"T","Υ":"Y","Χ":"X",
    "Δ":"D","Λ":"L","Σ":"S","Φ":"F","Θ":"Th","Ξ":"X","Ψ":"Ps","Ω":"O",
    "ο":"o","ρ":"p","κ":"k","ι":"i","ν":"v","χ":"x","τ":"t","μ":"m","υ":"y","σ":"s","ς":"s","φ":"f","ψ":"ps",
    # Cirílico
    "А":"A","В":"B","Е":"E","К":"K","М":"M","Н":"H","О":"O","Р":"P","С":"C","Т":"T","У":"Y","Х":"X","І":"I",
    "а":"a","е":"e","к":"k","м":"m","н":"h","о":"o","р":"p","с":"c","т":"t","у":"y","х":"x","і":"i",
}
TRANSL_TABLE = str.maketrans(HOMOGLYPH_MAP)

def looks_non_latin(s: Optional[str]) -> bool:
    if not s:
        return False
    for ch in s:
        cp = ord(ch)
        if 0x0370 <= cp <= 0x03FF or 0x0400 <= cp <= 0x04FF:
            return True
        if cp > 0x024F and not ch.isspace():
            return True
    return False

def fix_homoglyphs(s: Optional[str]) -> Optional[str]:
    if not s:
        return s
    return s.translate(TRANSL_TABLE)

# ----------------------------- Validações/Normalizações ----------------------
def cpf_is_valid(cpf: str) -> bool:
    cpf = clean_digits(cpf)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False
    def dig(digs: str) -> int:
        s = sum(int(d) * w for d, w in zip(digs, range(len(digs) + 1, 1, -1)))
        r = (s * 10) % 11
        return 0 if r == 10 else r
    return cpf[-2:] == f"{dig(cpf[:9])}{dig(cpf[:9] + str(dig(cpf[:9])))}"

def normalize_cpf(cpf: str) -> Optional[str]:
    d = clean_digits(cpf or "")
    if len(d) != 11 or not cpf_is_valid(d):
        return None
    return f"{d[0:3]}.{d[3:6]}.{d[6:9]}-{d[9:11]}"

def normalize_cnpj(cnpj: str) -> Optional[str]:
    d = clean_digits(cnpj or "")
    if len(d) != 14:
        return normalize_ws(cnpj) if cnpj else None
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"

def normalize_date(s: str) -> Optional[str]:
    if not s:
        return None
    try:
        s_clean = normalize_ws(s).replace("\\", "/").replace("-", "/").replace(".", "/")
        dt = dateparser.parse(s_clean, dayfirst=True, yearfirst=False)
        if not dt:
            return None
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return None

def normalize_sexo(val: str, norm_map: Dict[str, str]) -> Optional[str]:
    if not val:
        return None
    key = strip_accents_lower(val).strip()
    return norm_map.get(key, val.upper().strip())

def normalize_inclusao(val: str, norm_map: Dict[str, str]) -> Optional[str]:
    if not val:
        return None
    key = strip_accents_lower(val).replace(" ", "")
    if "24" in key:
        return norm_map.get("24h", "24h")
    return norm_map.get(key, normalize_ws(val).title())

def normalize_phone(ddd: Optional[str], fone: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    ddd_d = clean_digits(ddd) if ddd else None
    cel_d = clean_digits(fone) if fone else None

    if not (ddd_d and len(ddd_d) == 2):
        ddd_d = None

    if cel_d:
        if len(cel_d) in (8, 9):
            if len(cel_d) == 8:
                cel_d = "9" + cel_d
        elif len(cel_d) in (10, 11):
            if len(cel_d) == 10:
                cel_d = "9" + cel_d[2:]
            else:
                cel_d = cel_d[2:]

        if not ddd_d:
            ddd_d = cel_d[:2]

    if cel_d and len(cel_d) == 9:
        return ddd_d, f"{cel_d[:5]}-{cel_d[5:]}"
    return ddd_d, cel_d

def normalize_cep(val: Optional[str]) -> Optional[str]:
    if not val:
        return None
    d = clean_digits(val)
    if len(d) == 8:
        return f"{d[:5]}-{d[5:]}"
    return normalize_ws(val)

def normalize_pis(val: Optional[str]) -> Optional[str]:
    if not val:
        return None
    d = clean_digits(val)
    return d if len(d) == 11 else None

# ----------------------------- Config ----------------------------------------
@dataclass
class Config:
    output_sheet: str
    output_columns: Dict[str, str]
    required_fields: List[str]
    synonyms: Dict[str, List[str]]
    patterns: Dict[str, str]
    norm_sexo_map: Dict[str, str]
    norm_inclusao_map: Dict[str, str]

    @staticmethod
    def load(path: Path) -> "Config":
        with open(path, "r", encoding="utf-8") as f:
            y = yaml.safe_load(f)

        normalize = y.get("normalize", {})
        sexo_map = (normalize.get("sexo") or {}).get("map", {})
        incl_map = (normalize.get("beneficiario_inclusao") or {}).get("map", {})

        return Config(
            output_sheet=y["output_sheet"],
            output_columns=y["output_columns"],
            required_fields=y.get("required_fields", []),
            synonyms=y.get("synonyms", {}),
            patterns=y.get("patterns", {}),
            norm_sexo_map={strip_accents_lower(k): v for k, v in sexo_map.items()},
            norm_inclusao_map={strip_accents_lower(k): v for k, v in incl_map.items()},
        )

# ----------------------------- Heurísticas extras ----------------------------
LABEL_STOPWORDS = {
    "cpf","pis","cbo","nasc","nascimento","admiss","admissao","estado civil",
    "nacionalidade","sexo","mãe","mae","cep","matric","endereço","endereco",
    "lotacao","lotação","cnpj","grau","dependencia","dependência","acomodacao",
    "acomodação","inclusao","inclusão","email","e-mail","telefone","celular"
}

def _lines(text: str) -> List[str]:
    return [l.strip() for l in text.splitlines() if l.strip()]

def _extract_header_id_and_name(text: str) -> Tuple[Optional[str], Optional[str]]:
    for ln in _lines(text)[:60]:
        m = re.search(r"\b(\d{6,})\s*[-–—]\s*([A-ZÀ-Ü][A-ZÀ-Ü\s'.-]{4,})\b", ln)
        if m:
            _id = m.group(1)
            nome = re.sub(r"\s+", " ", m.group(2)).strip(" -–—.")
            return _id, nome
    return None, None

def _fallback_name_near_token(text: str, token_digits: str) -> Optional[str]:
    if not token_digits:
        return None
    L = _lines(text)
    idx = next((i for i, ln in enumerate(L) if token_digits in re.sub(r"\D+", "", ln)), None)
    if idx is None:
        return None
    for j in range(max(0, idx-3), min(len(L), idx+4)):
        if j == idx:
            continue
        cand = L[j]
        low = strip_accents_lower(cand)
        if any(sw in low for sw in LABEL_STOPWORDS):
            continue
        if re.search(r"\d", cand):
            continue
        if 2 <= len(cand.split()) <= 7 and 5 <= len(cand) <= 80:
            return re.sub(r"\s+", " ", cand).strip()
    return None

def _fallback_matricula(text: str) -> Optional[str]:
    for rx in [
        r"ficha[:.\s]*([A-Z0-9./\-]{2,20})",
        r"matr[íi]cula\s*[:\-]?\s*([A-Z0-9./\-]{2,20})",
        r"(?:registro|id\s*funcional|id)\s*[:\-]?\s*([A-Z0-9./\-]{2,20})",
        r"(?:c[oó]d\.?\s*funcional|codigo\s*funcional)\s*[:\-]?\s*([A-Z0-9./\-]{2,20})",
    ]:
        m = re.search(rx, text, flags=re.IGNORECASE | re.MULTILINE)
        if m:
            return m.group(1).strip()
    return None

def _sanitize_complemento(val: Optional[str]) -> Optional[str]:
    if not val:
        return None
    low = strip_accents_lower(val)
    if re.search(r"\b(salario|salário|motivo|classe|nivel|nível|aumento)\b", low):
        return None
    return normalize_ws(val)

# >>> Endereço e telefone (restritivo) ----------------------------------------
ADDRESS_LABELS = ["endereço", "endereco", "logradouro"]

def address_window(text: str, span: int = 240) -> Optional[str]:
    low = strip_accents_lower(text)
    for lab in ADDRESS_LABELS:
        i = low.find(lab)
        if i != -1:
            return text[i:i+span]
    return None

def extract_cep_and_numero_from_address(win: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not win:
        return None, None
    mcep = re.search(r"\b(\d{5})[-.\s]?(\d{3})\b", win)
    cep = f"{mcep.group(1)}-{mcep.group(2)}" if mcep else None
    numero = None
    wlow = strip_accents_lower(win)
    mrot = re.search(r"\b(?:n[ºo]|numero|número)\s*[:\-]?\s*(s/?n|\d{1,6})\b", wlow)
    if mrot:
        val = mrot.group(1)
        numero = "SN" if "s" in val else re.sub(r"\D", "", val)
    return cep, numero

def extract_phone_labeled(text: str, cfg: Config) -> Tuple[Optional[str], Optional[str]]:
    """
    Telefones só quando houver 'Celular' (palavra inteira) na linha.
    Ignora 'Telefone' para evitar pegar telefone da empresa.
    """
    labels = [*(cfg.synonyms.get("celular", []) or []), "celular", "cel."]
    word_labels = [rf"\b{re.escape(strip_accents_lower(lb))}\b" for lb in labels]

    for ln in text.splitlines():
        llow = strip_accents_lower(ln)
        if any(re.search(wl, llow) for wl in word_labels):
            m = re.search(r"(?:\(?(\d{2})\)?\s*)?([0-9]{4,5})[-.\s]?([0-9]{4})\b", ln)
            if m:
                ddd = m.group(1)
                num = (m.group(2) or "") + (m.group(3) or "")
                ddd_n, cel_n = normalize_phone(ddd, num)
                if ddd_n and not re.fullmatch(r"\d{2}", ddd_n):
                    ddd_n = None
                if cel_n and not re.fullmatch(r"\d{5}-\d{4}", cel_n):
                    cel_n = None
                return ddd_n, cel_n
    return None, None

# ----------------------------- Extração por página ---------------------------
def extract_text_from_pdf(pdf_path: Path) -> str:
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            pages_text = [(p.extract_text() or "") for p in pdf.pages]
            text = "\n".join(pages_text).strip()
    except Exception as e:
        LOG.error("Falha ao ler PDF %s: %s", pdf_path, e); return ""
    if len(text) >= 30:
        return text

    try:
        import shutil, subprocess, tempfile, os
        ocrmypdf = shutil.which(os.environ.get("OCRMYPDF_BIN", "ocrmypdf"))
        if not ocrmypdf:
            return text
        LOG.warning("Texto muito curto em %s (scan). Rodando OCR…", pdf_path.name)
        with tempfile.TemporaryDirectory() as td:
            ocr_pdf = Path(td) / "ocr.pdf"
            subprocess.run(
                [ocrmypdf, "--force-ocr", "--skip-text", "--output-type", "pdf", str(pdf_path), str(ocr_pdf)],
                check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE
            )
            with pdfplumber.open(str(ocr_pdf)) as pdf2:
                pages_text2 = [(p.extract_text() or "") for p in pdf2.pages]
                text2 = "\n".join(pages_text2).strip()
                if len(text2) > len(text):
                    return text2
    except Exception as e:
        LOG.warning("Falha no OCR para %s: %s", pdf_path.name, e)
    return text

def iter_page_texts(pdf_path: Path) -> List[str]:
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            texts = [(p.extract_text() or "").strip() for p in pdf.pages]
            if sum(len(t) for t in texts) >= 30:
                return texts
            return texts
    except Exception as e:
        LOG.error("Falha ao ler páginas do PDF %s: %s", pdf_path, e)
        return []

# ----------------------------- OCR por página --------------------------------
def ocr_single_page_text(pdf_path: Path, page_one_based: int, lang: str) -> Optional[str]:
    try:
        import shutil, subprocess, tempfile, os
        ocrmypdf = shutil.which(os.environ.get("OCRMYPDF_BIN", "ocrmypdf")) or shutil.which("ocrmypdf")
        if not ocrmypdf:
            return None
        with tempfile.TemporaryDirectory() as td:
            ocr_pdf = Path(td) / "ocr_one.pdf"
            cmd = [
                ocrmypdf, "-l", lang, "--force-ocr", "--skip-text", "--output-type", "pdf",
                "--pages", str(page_one_based), str(pdf_path), str(ocr_pdf)
            ]
            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            with pdfplumber.open(str(ocr_pdf)) as pdf2:
                if len(pdf2.pages) >= 1:
                    return (pdf2.pages[0].extract_text() or "").strip()
    except Exception as e:
        LOG.debug("Falha no OCR por página (ocrmypdf) %s p.%d: %s", pdf_path.name, page_one_based, e)
    return None

def ocr_single_page_text_tesseract(pdf_path: Path, page_one_based: int, lang: str) -> Optional[str]:
    try:
        from pdf2image import convert_from_path
        import pytesseract
        imgs = convert_from_path(str(pdf_path), dpi=300, first_page=page_one_based, last_page=page_one_based)
        if not imgs:
            return None
        txt = pytesseract.image_to_string(imgs[0], lang=lang)
        return (txt or "").strip()
    except Exception as e:
        LOG.debug("Falha no OCR por página (tesseract) %s p.%d: %s", pdf_path.name, page_one_based, e)
        return None

# ----------------------------- Busca por rótulo/padrão -----------------------
def window_after_label(text_norm: str, label: str, span: int = 120) -> Optional[str]:
    idx = text_norm.find(label)
    if idx == -1:
        return None
    return text_norm[idx + len(label): idx + len(label) + span]

def find_by_label_then_pattern(text: str, label_variants: List[str], pattern: str) -> Optional[str]:
    text_norm = strip_accents_lower(text)
    for label in label_variants:
        lab = strip_accents_lower(label)
        win = window_after_label(text_norm, lab, span=160)
        if win:
            m = re.search(pattern, win, flags=re.IGNORECASE)
            if m:
                return m.group(0)
    return None

def find_free_pattern(text: str, pattern: str) -> Optional[str]:
    return find_first(pattern, text)

def find_labeled_value(text: str, labels: List[str], maxlen: int = 96) -> Optional[str]:
    for lab in labels:
        rx = rf"{re.escape(lab)}\s*[:\-]?\s*(.+)"
        m = re.search(rx, text, flags=re.IGNORECASE)
        if m:
            raw = m.group(1).split("\n", 1)[0]
            raw = re.split(r"[;•|]", raw)[0].strip(" \t:.-")
            if raw:
                return raw[:maxlen]
    return None

# -------- Captura dedicada do valor após 'Mãe:' ------------------------------
def extract_mae_label_colon(text: str) -> Optional[str]:
    NBSP = "\u00A0"
    m = re.search(rf"(?im)^\s*M[ÃA]{NBSP}?E\s*:\s*([^\r\n]+)", text)
    if not m:
        m = re.search(r"(?i)\bM[ÃA]\s*E\s*:\s*([^\r\n]+)", text)
    if not m:
        return None
    val = m.group(1)
    val = re.split(
        r"\b(PAI|PIS|CPF|NACIONALIDADE|NATURALIDADE|DATA|ENDEREÇO|ENDERECO)\b\s*:",
        val, flags=re.IGNORECASE
    )[0]
    val = re.sub(r"\s{2,}", " ", val).replace(NBSP, " ").strip(" \t:.-")
    return val[:96] or None

# ----------------------------- Extração de campos ----------------------------
def extract_fields(text: str, cfg: Config) -> Dict[str, Optional[str]]:
    out: Dict[str, Optional[str]] = {k: None for k in cfg.output_columns.keys()}
    pat, syn = cfg.patterns, cfg.synonyms

    hdr_id, hdr_name = _extract_header_id_and_name(text)
    if hdr_id:
        out["titular_matricula"] = hdr_id
    if hdr_name:
        out["titular_nome"] = hdr_name
        out["beneficiario_nome"] = out.get("beneficiario_nome") or hdr_name

    if "cpf" in pat:
        out["cpf"] = (find_by_label_then_pattern(text, syn.get("cpf", ["cpf"]), pat["cpf"])
                      or find_free_pattern(text, pat["cpf"]))
    if "data" in pat:
        out["nascimento"] = find_by_label_then_pattern(text, syn.get("nascimento", []), pat["data"])
        out["data_admissao"] = find_by_label_then_pattern(text, syn.get("data_admissao", []), pat["data"])
    if "pis" in pat:
        out["pis"] = (find_by_label_then_pattern(text, syn.get("pis", []), pat["pis"])
                      or find_labeled_value(text, syn.get("pis", []), 64)
                      or find_free_pattern(text, pat["pis"]))
    if "cbo" in pat:
        out["cbo"] = (find_by_label_then_pattern(text, syn.get("cbo", []), pat["cbo"])
                      or find_free_pattern(text, pat["cbo"]))

    # CEP/Nº (Nº ficará em branco)
    addr_win = address_window(text)
    cep_from_addr, _numero_from_addr = extract_cep_and_numero_from_address(addr_win)
    out["cep"] = cep_from_addr

    if "email" in pat:
        out["email"] = (find_by_label_then_pattern(text, syn.get("email", []), pat["email"])
                        or find_free_pattern(text, pat["email"]))

    # Telefone apenas com 'Celular' na linha
    ddd_l, cel_l = extract_phone_labeled(text, cfg)
    out["ddd"], out["celular"] = ddd_l, cel_l

    text_norm = strip_accents_lower(text)

    def grab_after_label(key: str, maxlen: int = 64) -> Optional[str]:
        labels = syn.get(key, [])
        v = find_labeled_value(text, labels, maxlen)
        if v:
            return v
        for label in labels:
            lab = strip_accents_lower(label)
            win = window_after_label(text_norm, lab, span=maxlen)
            if win:
                raw = win.split("\n")[0]
                raw = re.split(r"[;•|]", raw)[0].strip(" :.-\t")
                if raw:
                    return raw
        return None

    out["beneficiario_nome"] = out.get("beneficiario_nome") or grab_after_label("beneficiario_nome", 96)
    out["beneficiario_grau_dependencia"] = grab_after_label("beneficiario_grau_dependencia", 48)

    # >>> SOS desativado
    out["sos_sn"] = None

    out["beneficiario_cnpj_lotacao"] = (
        find_by_label_then_pattern(text, syn.get("beneficiario_cnpj_lotacao", []), pat.get("cnpj", r".{0,18}"))
        or grab_after_label("beneficiario_cnpj_lotacao", 32)
    )
    out["beneficiario_tipo_acomodacao"] = grab_after_label("beneficiario_tipo_acomodacao", 32)

    incl = grab_after_label("beneficiario_inclusao", 32)
    if incl:
        m = re.search(r"\b(?:24\s*h|24h|programada|agendada)\b", strip_accents_lower(incl))
        out["beneficiario_inclusao"] = m.group(0) if m else None

    out["nacionalidade"] = grab_after_label("nacionalidade", 32)
    out["estado_civil"] = grab_after_label("estado_civil", 24)
    out["sexo"] = grab_after_label("sexo", 8)

    mae_colon = extract_mae_label_colon(text)
    out["mae_nome"] = mae_colon if mae_colon else grab_after_label("mae_nome", 96)

    out["titular_nome"] = out.get("titular_nome") or grab_after_label("titular_nome", 96)
    out["titular_matricula"] = out.get("titular_matricula") or grab_after_label("titular_matricula", 32) or _fallback_matricula(text)

    # Força número do endereço em branco (empresa ≠ beneficiário)
    out["numero"] = None
    out["complemento"] = _sanitize_complemento(grab_after_label("complemento", 64))

    cpf_digits = clean_digits(out.get("cpf") or "")
    if not out.get("beneficiario_nome"):
        out["beneficiario_nome"] = _fallback_name_near_token(text, cpf_digits) or out.get("beneficiario_nome")
    if not out.get("titular_nome") and out.get("beneficiario_nome"):
        out["titular_nome"] = out["beneficiario_nome"]
    if not out.get("beneficiario_nome") and out.get("titular_nome"):
        out["beneficiario_nome"] = out["titular_nome"]

    return out

# --- Fix de nomes com OCR por página ----------------------------------------
def fix_names_if_needed(
    row: Dict[str, Optional[str]],
    pdf_path: Path,
    page_idx_one_based: int,
    cfg: Config,
    original_page_text: str,
    force_ocr: bool,
    ocr_lang: str
) -> Dict[str, Optional[str]]:
    keys = ("mae_nome", "beneficiario_nome", "titular_nome")
    need_ocr = force_ocr or any(looks_non_latin(row.get(k) or "") for k in keys)
    if not need_ocr:
        for k in keys:
            row[k] = fix_homoglyphs(row.get(k))
        return row

    ocr_text = ocr_single_page_text(pdf_path, page_idx_one_based, ocr_lang)
    if not ocr_text:
        ocr_text = ocr_single_page_text_tesseract(pdf_path, page_idx_one_based, ocr_lang)
    if not ocr_text:
        LOG.warning("Sem backend de OCR disponível para %s p.%d. Instale 'ocrmypdf' OU 'tesseract'+'pdf2image'.",
                    pdf_path.name, page_idx_one_based)
        for k in keys:
            row[k] = fix_homoglyphs(row.get(k))
        return row

    re_row = extract_fields(ocr_text, cfg)
    for k in keys:
        if re_row.get(k):
            row[k] = re_row[k]
    return row

# ----------------------------- Normalização ----------------------------------
def clean_and_normalize(row: Dict[str, Optional[str]], cfg: Config) -> Tuple[Dict[str, Optional[str]], List[str]]:
    issues: List[str] = []

    for nm in ("beneficiario_nome","titular_nome","mae_nome"):
        row[nm] = fix_homoglyphs(row.get(nm))

    row["cpf"] = normalize_cpf(row.get("cpf") or "")
    if not row["cpf"]:
        issues.append("CPF inválido/ausente")

    for k in ("nascimento", "data_admissao"):
        if row.get(k):
            row[k] = normalize_date(row[k])
        if not row.get(k):
            issues.append(f"Data inválida/ausente: {k}")

    if row.get("sexo"):
        row["sexo"] = normalize_sexo(row["sexo"], cfg.norm_sexo_map)

    if row.get("beneficiario_inclusao"):
        row["beneficiario_inclusao"] = normalize_inclusao(row["beneficiario_inclusao"], cfg.norm_inclusao_map)

    if row.get("beneficiario_cnpj_lotacao"):
        row["beneficiario_cnpj_lotacao"] = normalize_cnpj(row["beneficiario_cnpj_lotacao"])

    if row.get("cep"):
        row["cep"] = normalize_cep(row["cep"])

    if row.get("pis"):
        row["pis"] = normalize_pis(row["pis"])

    row["ddd"], row["celular"] = normalize_phone(row.get("ddd"), row.get("celular"))

    if row.get("ddd") and not re.fullmatch(r"\d{2}", row["ddd"] or ""):
        row["ddd"] = None
    if row.get("celular") and not re.fullmatch(r"\d{5}-\d{4}", row["celular"] or ""):
        row["celular"] = None

    for k, v in list(row.items()):
        if isinstance(v, str):
            row[k] = normalize_ws(v)

    for k in cfg.required_fields:
        if not (row.get(k) and str(row[k]).strip()):
            issues.append(f"Campo obrigatório ausente: {k}")

    return row, issues

# ----------------------------- Cabeçalhos / Auditoria ------------------------
def _detect_header_map(ws, expected_headers, max_scan=30):
    best = None
    for r in range(1, max_scan + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_set = set([str(v).strip().lower() for v in vals if v is not None])
        matches = sum(1 for h in expected_headers if h.lower() in row_set)
        if matches and (best is None or matches > best[1]):
            best = (r, matches, vals)
    if not best:
        raise RuntimeError("Não encontrei a linha de cabeçalho na aba do modelo.")
    header_row, _, vals = best
    header_map = {}
    for c, v in enumerate(vals, start=1):
        if v is not None:
            header_map[str(v).strip()] = c
    return header_row, header_map

def _ensure_template_headers(cfg, template_path: Path):
    wb = load_workbook(template_path, data_only=False)
    try:
        ws = wb[cfg.output_sheet] if cfg.output_sheet in wb.sheetnames else wb.active
        expected_headers = list(cfg.output_columns.values())
        header_row, header_map = _detect_header_map(ws, expected_headers, max_scan=30)
        key_to_col, missing = {}, []
        for key, col_name in cfg.output_columns.items():
            col = header_map.get(col_name)
            if col is None:
                missing.append(col_name)
            else:
                key_to_col[key] = col
        if missing:
            raise RuntimeError("Cabeçalhos do YAML não encontrados no modelo:\n - " + "\n - ".join(missing))
        return header_row, header_map, key_to_col
    finally:
        try:
            wb.close()
        except Exception:
            pass

def write_column_audit(audit_info: dict, cfg: Config, outdir: Path):
    outdir.mkdir(parents=True, exist_ok=True)
    path = Path(outdir) / "column_map.csv"
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["key", "header_name", "col_number"])
        for key in cfg.output_columns.keys():
            header_name = cfg.output_columns[key]
            col = audit_info["key_to_col"].get(key)
            w.writerow([key, header_name, col if col is not None else "MISSING"])
    LOG.info("Mapa de colunas salvo em: %s", path)

# ----------------------------- Escrita: layout idêntico (append) -------------
TEXT_KEYS = {
    "beneficiario_cnpj_lotacao",
    "cpf", "pis", "cep", "ddd", "celular", "titular_matricula", "numero", "cbo",
    "nascimento", "data_admissao"
}

def write_to_xlsx_exact_layout(rows, cfg, out_xlsx, template_path, write_markers=False, fresh=False):
    from pathlib import Path
    import shutil
    from openpyxl import load_workbook

    template_path = Path(template_path) if template_path else None
    if not (template_path and template_path.exists()):
        raise RuntimeError(
            "Para manter layout idêntico, informe --template apontando para o modelo de planilha base (XLSX)."
        )

    expected_headers = list(cfg.output_columns.values())

    wb_model = load_workbook(template_path, data_only=False)
    try:
        ws_model = wb_model[cfg.output_sheet] if cfg.output_sheet in wb_model.sheetnames else wb_model.active
        header_row, header_map = _detect_header_map(ws_model, expected_headers, max_scan=30)
    finally:
        try:
            wb_model.close()
        except Exception:
            pass

    # ===== tenta COM (mantém validações) =====
    try:
        import win32com.client as win32
        out_xlsx = Path(out_xlsx); out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        use_existing = out_xlsx.exists() and not fresh
        if not use_existing:
            shutil.copyfile(str(template_path), str(out_xlsx))

        excel = win32.Dispatch("Excel.Application"); excel.Visible = False
        wb_com = excel.Workbooks.Open(str(out_xlsx.resolve()))
        ws_com = wb_com.Worksheets(cfg.output_sheet)

        xlUp = -4162
        xlLeft = -4131
        xlColorIndexNone = -4142  # "Sem preenchimento"

        last_row = ws_com.Cells(ws_com.Rows.Count, 1).End(xlUp).Row
        row_idx = max(last_row, header_row) + 1

        col_index = {name: header_map[name] for name in cfg.output_columns.values() if name in header_map}

        for r in rows:
            r_sane = {k: (normalize_ws(v) if isinstance(v, str) else v) for k, v in r.items()}

            # escrevendo valores e formatando
            for key, col_name in cfg.output_columns.items():
                col = col_index.get(col_name)
                if not col:
                    continue
                val = r_sane.get(key, "") or ""

                # filtros finais p/ telefone
                if key == "ddd" and (not re.fullmatch(r"\d{2}", str(val))):
                    val = ""
                if key == "celular" and (val and not re.fullmatch(r"\d{5}-\d{4}", str(val))):
                    val = ""

                cell = ws_com.Cells(row_idx, col)
                if key in TEXT_KEYS:
                    cell.NumberFormat = "@"
                cell.HorizontalAlignment = xlLeft
                cell.IndentLevel = 0
                try:
                    cell.Font.Name = "Calibri"; cell.Font.Size = 11
                except Exception:
                    pass

                cell.Value = str(val) if key in TEXT_KEYS else val

            # pinta de amarelo somente as CÉLULAS VAZIAS desta linha
            for col in col_index.values():
                c = ws_com.Cells(row_idx, col)
                v = c.Value
                is_empty = (v is None) or (str(v).strip() == "")
                try:
                    if is_empty:
                        c.Interior.ColorIndex = 36  # amarelo claro
                    else:
                        c.Interior.ColorIndex = xlColorIndexNone  # remove qualquer preenchimento
                except Exception:
                    if is_empty:
                        c.Interior.ColorIndex = 6    # fallback amarelo

            row_idx += 1

        wb_com.Save(); wb_com.Close(SaveChanges=True); excel.Quit()
        return {
            "header_row": header_row,
            "header_map": header_map,
            "key_to_col": {k: col_index[cfg.output_columns[k]] for k in cfg.output_columns if cfg.output_columns[k] in col_index}
        }
    except Exception as e:
        LOG.warning("Falha Excel COM (pywin32). Usando openpyxl. Erro: %s", e)

    # ===== Fallback openpyxl =====
    out_xlsx = Path(out_xlsx); out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    use_existing = out_xlsx.exists() and not fresh
    wb2 = load_workbook(out_xlsx if use_existing else template_path, data_only=False)
    try:
        ws2 = wb2[cfg.output_sheet] if cfg.output_sheet in wb2.sheetnames else wb2.active
        header_row2, header_map2 = _detect_header_map(ws2, expected_headers, max_scan=30)
        key_to_col2 = {k: header_map2[cfg.output_columns[k]] for k in cfg.output_columns if cfg.output_columns[k] in header_map2}

        row_idx = header_row2 + 1

        def row_has_data(r):
            return any(ws2.cell(row=r, column=c).value not in (None, "") for c in range(1, ws2.max_column + 1))
        while row_has_data(row_idx):
            row_idx += 1

        calibri = Font(name="Calibri", size=11)
        yellow = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")

        for r in rows:
            r_sane = {k: (normalize_ws(v) if isinstance(v, str) else v) for k, v in r.items()}

            # escreve valores na linha
            for k, col in key_to_col2.items():
                val = r_sane.get(k, "") or ""
                if k == "ddd" and (not re.fullmatch(r"\d{2}", str(val))):
                    val = ""
                if k == "celular" and (val and not re.fullmatch(r"\d{5}-\d{4}", str(val))):
                    val = ""

                cell = ws2.cell(row=row_idx, column=col)
                cell.value = val
                cell.font = calibri
                cell.alignment = Alignment(horizontal="left", indent=0)
                if k in TEXT_KEYS:
                    cell.number_format = "@@" if False else "@"
                    # (mantido igual, mas você pode tirar esse if bobo se quiser)

            # pinta de amarelo somente as CÉLULAS VAZIAS desta linha
            for col in key_to_col2.values():
                cell = ws2.cell(row=row_idx, column=col)
                v = cell.value
                is_empty = (v is None) or (str(v).strip() == "")
                if is_empty:
                    cell.fill = yellow
                else:
                    cell.fill = PatternFill(fill_type=None)  # limpa

            row_idx += 1

        try:
            wb2.save(out_xlsx)
        except PermissionError as e:
            raise RuntimeError(
                f"Permissão negada ao salvar '{out_xlsx}'. "
                f"Feche o arquivo no Excel (se estiver aberto) ou use --fresh / --xlsx-name para outro caminho."
            ) from e
    finally:
        try:
            wb2.close()
        except Exception:
            pass

    return {"header_row": header_row2, "header_map": header_map2, "key_to_col": key_to_col2}

# ----------------------------- Escrita (fallback simples) --------------------
def write_to_xlsx(rows: List[Dict[str, Optional[str]]], cfg: Config, out_xlsx: Path, template: Optional[Path] = None):
    if template and template.exists():
        wb = load_workbook(template)
        ws = wb[cfg.output_sheet] if cfg.output_sheet in wb.sheetnames else wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = cfg.output_sheet
        ws.append([cfg.output_columns[k] for k in cfg.output_columns.keys()])

    keys_order = list(cfg.output_columns.keys())
    calibri = Font(name="Calibri", size=11)
    yellow = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")

    for row in rows:
        r_sane = [normalize_ws(row.get(k, "") or "") for k in keys_order]
        ws.append(r_sane)
        current_row = ws.max_row

        # formatação por coluna + filtros DDD/Celular
        for idx, key in enumerate(keys_order, start=1):
            cell = ws.cell(row=current_row, column=idx)

            if key == "ddd" and (not re.fullmatch(r"\d{2}", str(cell.value or ""))):
                cell.value = ""
            if key == "celular" and (cell.value and not re.fullmatch(r"\d{5}-\d{4}", str(cell.value))):
                cell.value = ""

            cell.font = calibri
            cell.alignment = Alignment(horizontal="left", indent=0)
            if key in TEXT_KEYS:
                cell.number_format = "@"

            # pinta amarelo apenas se esta célula estiver vazia
            v = cell.value
            is_empty = (v is None) or (str(v).strip() == "")
            if is_empty:
                cell.fill = yellow
            else:
                cell.fill = PatternFill(fill_type=None)

    for idx, head in enumerate([cfg.output_columns[k] for k in keys_order], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(len(head) + 4, 36))

    out_xlsx = Path(out_xlsx); out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    try:
        wb.close()
    except Exception:
        pass

# ----------------------------- Relatórios ------------------------------------
def write_reports(rows_ok: List[Dict[str, Optional[str]]], rows_err: List[Tuple[Dict[str, Optional[str]], List[str]]], outdir: Path):
    outdir.mkdir(parents=True, exist_ok=True)
    with open(outdir / "report.jsonl", "w", encoding="utf-8") as fj:
        for r in rows_ok:
            fj.write(json.dumps({"ok": True, "row": r}, ensure_ascii=False) + "\n")
        for r, errs in rows_err:
            fj.write(json.dumps({"ok": False, "row": r, "errors": errs}, ensure_ascii=False) + "\n")

    rej_cols = ["erro", *list(rows_err[0][0].keys())] if rows_err else ["erro"]
    with open(outdir / "rejeitados.csv", "w", encoding="utf-8", newline="") as fc:
        w = csv.writer(fc, delimiter=";"); w.writerow(rej_cols)
        for r, errs in rows_err:
            w.writerow([" | ".join(errs), *[r.get(k, "") for k in (rej_cols[1:] if len(rej_cols) > 1 else [])]])

# ----------------------------- Dedupe (XLSX existente) -----------------------
def read_existing_keys_from_xlsx(path: Path, cfg: Config) -> set[tuple[str, str]]:
    if not path.exists():
        return set()
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[cfg.output_sheet] if cfg.output_sheet in wb.sheetnames else wb.active
        expected_headers = list(cfg.output_columns.values())
        header_row, header_map = _detect_header_map(ws, expected_headers, max_scan=30)
        col_cpf = header_map.get(cfg.output_columns["cpf"])
        col_mat = header_map.get(cfg.output_columns["titular_matricula"])
        if not (col_cpf and col_mat):
            return set()
        seen = set()
        for r in range(header_row + 1, ws.max_row + 1):
            vcpf = ws.cell(row=r, column=col_cpf).value
            vmat = ws.cell(row=r, column=col_mat).value
            cpf_d = clean_digits(str(vcpf or ""))
            mat = (str(vmat or "").strip())
            if cpf_d and mat:
                seen.add((cpf_d, mat))
        return seen
    finally:
        try:
            wb.close()
        except Exception:
            pass

# ----------------------------- Coleta PDFs -----------------------------------
def collect_pdf_paths(inputs: List[str]) -> List[Path]:
    paths: List[Path] = []
    for inp in inputs:
        p = Path(inp)
        if p.is_dir():
            paths.extend(sorted(p.rglob("*.pdf")))
        elif any(ch in inp for ch in "*?[]"):
            paths.extend([Path(x) for x in glob(inp)])
        else:
            if p.suffix.lower() == ".pdf":
                paths.append(p)
    return [x for x in paths if x.exists()]

# ----------------------------- Main ------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="PDF→Excel Ingestor (MVP)")
    ap.add_argument(
        "-i", "--input",
        nargs="+",
        default=[str(ROOT / "entrada")],
        help="Arquivos PDF, pastas ou glob (*.pdf)"
    )
    ap.add_argument(
        "-m", "--mapping",
        default=str(ROOT / "mapping.yaml"),
        help="YAML de configuração/mapeamento"
    )
    ap.add_argument(
        "-t", "--template",
        default=str(ROOT / "MODELO_PLANILHA_INCLUSAO.xlsx"),
        help="Modelo XLSX da planilha base (obrigatório p/ layout idêntico)"
    )
    ap.add_argument(
        "-o", "--outdir",
        default=str(ROOT / "saida"),
        help="Diretório de saída (XLSX e relatórios)"
    )
    ap.add_argument(
        "--xlsx-name",
        default="inclusao_beneficiarios.xlsx",
        help="Nome do XLSX de saída"
    )
    ap.add_argument("--loglevel", default="INFO", help="Nível de log (DEBUG, INFO, WARNING, ERROR)")
    ap.add_argument("--check-template", action="store_true", help="Só valida o template (cabeçalhos) e sai")
    ap.add_argument("--audit-columns", action="store_true", help="Gera saida/column_map.csv com o mapeamento coluna->campo")
    ap.add_argument("--write-markers", action="store_true", help="Escreve uma linha de marcadores <<key>> antes dos dados")
    ap.add_argument("--write-even-if-errors", action="store_true", help="Também escreve no Excel as linhas com erro (para conferência).")
    ap.add_argument("--dump-text", action="store_true", help="Salva o texto extraído por página em saida/trace/<pdf>_pNNN.txt")
    ap.add_argument("--relax-required", action="store_true", help="Não rejeita linhas por obrigatórios ausentes; apenas avisa.")
    ap.add_argument("--fresh", action="store_true", help="Recria o XLSX a partir do template (não acumula).")
    ap.add_argument("--no-reports", action="store_true", help="Não gerar reports.")
    ap.add_argument("--force-ocr", action="store_true", help="Força OCR em todas as páginas.")
    ap.add_argument("--ocr-lang", default="por+eng", help="Idiomas para OCR (tesseract/ocrmypdf), ex.: 'por+eng'.")

    args = ap.parse_args()
    logging.basicConfig(level=getattr(logging, args.loglevel.upper(), logging.INFO), format=LOG_FORMAT)

    for maybe_dir in args.input:
        p = Path(maybe_dir)
        if p.is_dir():
            p.mkdir(parents=True, exist_ok=True)

    outdir = Path(args.outdir); outdir.mkdir(parents=True, exist_ok=True)
    xlsx_path = outdir / args.xlsx_name
    template_path = Path(args.template) if args.template else None
    cfg = Config.load(Path(args.mapping))

    if args.check_template:
        header_row, header_map, key_to_col = _ensure_template_headers(cfg, Path(args.template))
        LOG.info("Template OK. Aba '%s' | Cabeçalho na linha %d | %d colunas mapeadas.", cfg.output_sheet, header_row, len(key_to_col))
        if args.audit_columns and not args.no_reports:
            write_column_audit({"key_to_col": key_to_col}, cfg, Path(args.outdir))
        return 0

    pdfs = collect_pdf_paths(args.input)
    if not pdfs:
        LOG.error("Nenhum PDF encontrado em: %s", args.input)
        LOG.info("Dica: coloque arquivos .pdf dentro de %s", args.input)
        sys.exit(2)

    existing_keys: set[tuple[str, str]] = set()
    try:
        existing_keys = read_existing_keys_from_xlsx(xlsx_path, cfg)
        if existing_keys:
            LOG.info("Dedup: %d registro(s) já presentes no XLSX.", len(existing_keys))
    except Exception as e:
        LOG.warning("Não foi possível ler o XLSX existente para dedupe: %s", e)

    LOG.info("Processando %d PDF(s)...", len(pdfs))
    rows_ok: List[Dict[str, Optional[str]]] = []
    rows_err: List[Tuple[Dict[str, Optional[str]], List[str]]] = []
    seen_in_run: set[tuple[str, str]] = set()

    for pdf in pdfs:
        LOG.info("PDF: %s", pdf.name)
        page_texts = iter_page_texts(pdf)
        if not page_texts:
            LOG.warning("Sem texto extraído de %s", pdf.name)
            continue

        total_pages = len(page_texts)
        for idx, text in enumerate(page_texts, start=1):
            LOG.debug("Processando página %d/%d de %s", idx, total_pages, pdf.name)

            if args.dump_text:
                trace_dir = Path(args.outdir) / "trace"
                trace_dir.mkdir(parents=True, exist_ok=True)
                with open(trace_dir / f"{pdf.stem}_p{idx:03}.txt", "w", encoding="utf-8") as f:
                    f.write(text)

            if not text.strip():
                LOG.warning("Página %d sem texto extraído (%s)", idx, pdf.name)
                continue

            raw_row = extract_fields(text, cfg)
            raw_row = fix_names_if_needed(raw_row, pdf, idx, cfg, text, args.force_ocr, args.ocr_lang)
            row, issues = clean_and_normalize(raw_row, cfg)

            if args.relax_required:
                before = len(issues)
                issues = [e for e in issues if not e.startswith("Campo obrigatório ausente")]
                if before - len(issues):
                    LOG.debug("Relaxed required-field issue(s) para %s p.%d.", pdf.name, idx)

            cpf_d = clean_digits(row.get("cpf") or "")
            mat = (row.get("titular_matricula") or "").strip()
            key = (cpf_d, mat)

            if cpf_d and mat:
                if key in existing_keys:
                    LOG.info("Pág. %d pulada (duplicado no XLSX): CPF=%s | Matrícula=%s", idx, cpf_d, mat)
                    continue
                if key in seen_in_run:
                    issues.append("Duplicado (CPF+Matrícula)")

            if issues:
                rows_err.append((row, issues))
                LOG.warning("Erros em %s p.%d: %s", pdf.name, idx, " | ".join(issues))
            else:
                rows_ok.append(row)
                if cpf_d and mat:
                    seen_in_run.add(key)

    rows_to_write = rows_ok if not args.write_even_if_errors else (rows_ok + [r for r, _ in rows_err])

    audit_info = write_to_xlsx_exact_layout(
        rows_to_write, cfg, xlsx_path, template_path,
        write_markers=args.write_markers, fresh=args.fresh
    )

    if args.audit_columns and not args.no_reports:
        write_column_audit(audit_info, cfg, outdir)

    if not args.no_reports:
        write_reports(rows_ok, rows_err, outdir)

    LOG.info("✅ Concluído. Gerado: %s", xlsx_path)
    if not args.no_reports:
        LOG.info("Relatórios: %s / %s", outdir / "report.jsonl" , outdir / "rejeitados.csv")
    if rows_err:
        LOG.info("Linhas OK: %d | Rejeitadas: %d", len(rows_ok), len(rows_err))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())